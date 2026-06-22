import os
import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def create_excel_openpyxl():
    input_path = 'parsed_output.json'
    output_path = r'C:\Users\Cesar\Documents\GRUPO EMPRESARIAL REYES\PROYECTOS\RRSS_objetivo\Estrategia_Posicionamiento_Mes1.xlsx'
    
    with open(input_path, 'r', encoding='utf-8') as f:
        data = json.load(f)
        
    wb = openpyxl.Workbook()
    # Remover hoja default
    default_sheet = wb.active
    
    days_order = ['Lunes', 'Martes', 'Miércoles', 'Jueves', 'Viernes', 'Sábado', 'Domingo']
    
    # Styles
    border_side = Side(border_style="thin", color="000000")
    thin_border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
    
    header_fill = PatternFill("solid", fgColor="333333")
    header_font = Font(bold=True, color="FFFFFF")
    
    gen_fill = PatternFill("solid", fgColor="FF9900")
    pr_fill = PatternFill("solid", fgColor="1155cc")
    aq_fill = PatternFill("solid", fgColor="38761d")
    
    for day in days_order:
        if day not in data: continue
        ws = wb.create_sheet(title=day)
        
        # Fila 1: Headers super/Madres
        ws.merge_cells('A1:B1')
        ws['A1'] = "GENERAL"
        ws['A1'].fill = header_fill
        ws['A1'].font = header_font
        ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
        ws['A1'].border = thin_border
        ws['B1'].border = thin_border
        
        ws.merge_cells('C1:D1')
        ws['C1'] = "PRIMER REPORTE"
        ws['C1'].fill = pr_fill
        ws['C1'].font = header_font
        ws['C1'].alignment = Alignment(horizontal="center", vertical="center")
        ws['C1'].border = thin_border
        ws['D1'].border = thin_border
        
        ws.merge_cells('E1:F1')
        ws['E1'] = "ACTIVA QR"
        ws['E1'].fill = aq_fill
        ws['E1'].font = header_font
        ws['E1'].alignment = Alignment(horizontal="center", vertical="center")
        ws['E1'].border = thin_border
        ws['F1'].border = thin_border
        
        ws.merge_cells('G1:H1')
        ws['G1'] = "AUDIOVISUAL / RESULTADOS"
        ws['G1'].fill = gen_fill
        ws['G1'].font = header_font
        ws['G1'].alignment = Alignment(horizontal="center", vertical="center")
        ws['G1'].border = thin_border
        ws['H1'].border = thin_border
        
        # Fila 2: Nombres exactos de columnas
        columns = [
            "Semana", "Objetivo del Día", 
            "Post Mañana", "Historia Noche",
            "Post Mañana", "Historia Noche",
            "Reel / Video César", "Sinergia / Acciones Clave"
        ]
        
        key_mapping = [
            "Semana", "Objetivo del Día",
            "Estrategia Primer Reporte (Post Mañana)", "Estrategia Primer Reporte (Historia Noche)",
            "Estrategia ActivaQR (Post Mañana)", "Estrategia ActivaQR (Historia Noche)",
            "Reel / Video César (Si Aplica)", "Sinergia y Acciones Clave"
        ]
        
        for col_idx, col_name in enumerate(columns, start=1):
            cell = ws.cell(row=2, column=col_idx)
            cell.value = col_name
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = thin_border
            
        # Rellenar con los datos
        current_row = 3
        for item in data[day]:
            for col_idx, key in enumerate(key_mapping, start=1):
                cell = ws.cell(row=current_row, column=col_idx)
                # Remueve las dobles diagonales y formatéalo como texto limpio
                val = item[key]
                if isinstance(val, str):
                    val = val.replace('\\#', '#').replace('\\[', '[').replace('\\]', ']')
                cell.value = val
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                cell.border = thin_border
                if col_idx == 1:
                    cell.font = Font(bold=True)
            current_row += 1
            
        # Set column widths
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 45
        ws.column_dimensions['D'].width = 40
        ws.column_dimensions['E'].width = 45
        ws.column_dimensions['F'].width = 40
        ws.column_dimensions['G'].width = 45
        ws.column_dimensions['H'].width = 45
        
        # Set row heights
        ws.row_dimensions[1].height = 25
        ws.row_dimensions[2].height = 25
        for r in range(3, current_row):
            # Aumentamos el tamaño porque ahora hay textos integrales largos
            ws.row_dimensions[r].height = 250
            
    # Eliminar Default
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']
        
    wb.save(output_path)
    print(f"Excel generado utilizando base JSON integra en: {output_path}")

create_excel_openpyxl()
