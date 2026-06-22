"""
Generador CORRECTO de Calendario Excel 2026
Fuente: BLOG_ESTRATEGICO_2026.csv
- 7 hojas = días de la semana (LUNES a DOMINGO)
- 24 artículos distribuidos en ciclo de 4 semanas x 6 días activos
- Cada artículo tiene su copy de FB, IG, LI y X
"""

import csv
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─── PALETA ────────────────────────────────────────────────────────────────
C_ROJO        = "C0392B"
C_BLANCO      = "FFFFFF"
C_NEGRO       = "0D0D0D"
C_GRIS_DARK   = "2C3E50"
C_AMARILLO    = "F39C12"
C_AZUL_DARK   = "154360"
C_FB          = "1877F2"   # Facebook azul
C_IG          = "833AB4"   # Instagram púrpura
C_LI          = "0A66C2"   # LinkedIn azul
C_X           = "14171A"   # X negro
C_VERDE       = "1E8449"

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def font(bold=False, size=10, color="0D0D0D", italic=False, name="Calibri"):
    return Font(bold=bold, size=size, color=color, italic=italic, name=name)

def border():
    thin = Side(border_style="thin", color="CCCCCC")
    return Border(left=thin, right=thin, top=thin, bottom=thin)

def center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def left():
    return Alignment(horizontal="left", vertical="top", wrap_text=True)

# ─── LEER CSV ──────────────────────────────────────────────────────────────
csv_path = r"C:\Users\Cesar\Documents\GRUPO EMPRESARIAL REYES\PROYECTOS\RRSS_objetivo\BLOG_ESTRATEGICO_2026.csv"
articles = []
with open(csv_path, encoding="utf-8-sig", newline="") as f:
    reader = csv.DictReader(f)
    for row in reader:
        if row.get("ID"):
            articles.append(row)

# ─── MAPEO: Día de la semana para los 24 artículos ─────────────────────────
# Estrategia: 4 semanas × 6 artículos por semana (L-S), domingo = descanso/especial
# Artículos 1-6 = Semana 1, 7-12 = Semana 2, 13-18 = Semana 3, 19-24 = Semana 4
DIAS = ["LUNES", "MARTES", "MIÉRCOLES", "JUEVES", "VIERNES", "SÁBADO"]
SEMANAS = [
    {"num": 1, "titulo": "El Mito de la Clientela Fija", "articulos": articles[0:6]},
    {"num": 2, "titulo": "El Error del Sobrinitis",       "articulos": articles[6:12]},
    {"num": 3, "titulo": "Fricción Operativa (PDF vs Link)", "articulos": articles[12:18]},
    {"num": 4, "titulo": "El Techo del Excel (Sistemas)", "articulos": articles[18:24]},
]

# Construir mapa DÍA → lista de (semana, artículo)
# día index 0=Lunes, 1=Martes... 5=Sábado
dia_map = {dia: [] for dia in DIAS}
for sem in SEMANAS:
    for i, art in enumerate(sem["articulos"]):
        dia = DIAS[i]  # Cada artículo va al día correspondiente
        dia_map[dia].append((sem["num"], sem["titulo"], art))

# ─── CREAR WORKBOOK ────────────────────────────────────────────────────────
wb = openpyxl.Workbook()
wb.remove(wb.active)

ORDEN_DIAS = ["LUNES", "MARTES", "MIÉRCOLES", "JUEVES", "VIERNES", "SÁBADO", "DOMINGO"]

# Color de encabezado por red
RED_CONFIG = {
    "Facebook":  {"fill": fill(C_FB),      "icon": "📘"},
    "Instagram": {"fill": fill(C_IG),      "icon": "📸"},
    "LinkedIn":  {"fill": fill(C_LI),      "icon": "💼"},
    "X (Twitter)": {"fill": fill(C_X),    "icon": "🐦"},
}

for dia in ORDEN_DIAS:
    ws = wb.create_sheet(title=dia.capitalize())
    ws.sheet_view.showGridLines = False

    entries = dia_map.get(dia, [])

    # ── TÍTULO DE LA HOJA ──────────────────────────────────────────────────
    ncols = 11  # Total de columnas
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    t = ws.cell(row=1, column=1, value=f"📅 {dia}  —  Calendario de Publicaciones RRSS 2026")
    t.fill = fill(C_ROJO)
    t.font = Font(bold=True, size=13, color=C_BLANCO, name="Calibri")
    t.alignment = center()
    ws.row_dimensions[1].height = 26

    # ── CABECERA DE COLUMNAS ───────────────────────────────────────────────
    HEADERS = [
        ("A", "Semana",            12,  C_GRIS_DARK),
        ("B", "ID Artículo",       10,  C_GRIS_DARK),
        ("C", "Título del Blog",   40,  C_GRIS_DARK),
        ("D", "Silo / Categoría",  20,  C_GRIS_DARK),
        ("E", "Tipo de Post",      16,  C_GRIS_DARK),
        ("F", "📘 Copy Facebook",  55,  C_FB),
        ("G", "📸 Copy Instagram", 55,  C_IG),
        ("H", "💼 Copy LinkedIn",  55,  C_LI),
        ("I", "🐦 Copy X (Twitter)",55, C_X),
        ("J", "🖼️ URL Imagen",     45,  "346855"),
        ("K", "🔗 URL Artículo",   50,  "154360"),
    ]

    for col_idx, (col_letter, header, width, color) in enumerate(HEADERS, start=1):
        h = ws.cell(row=2, column=col_idx, value=header)
        h.fill = fill(color)
        h.font = Font(bold=True, size=9, color=C_BLANCO, name="Calibri")
        h.alignment = center()
        h.border = border()
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[2].height = 30

    # ── FILAS DE DATOS ─────────────────────────────────────────────────────
    current_row = 3

    if not entries:
        # DOMINGO: sin contenido programado
        ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=ncols)
        c = ws.cell(row=3, column=1,
                    value="🌿 Domingo de Descanso — Sin publicaciones programadas. Día de análisis, planificación y carga de contenidos para la semana.")
        c.fill = fill("F4F6F7")
        c.font = Font(size=10, italic=True, color="566573", name="Calibri")
        c.alignment = center()
        ws.row_dimensions[3].height = 30
    else:
        for sem_num, sem_titulo, art in entries:
            # Fila de separación por semana
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=ncols)
            sc = ws.cell(
                row=current_row, column=1,
                value=f"  📌 SEMANA {sem_num}: {sem_titulo}"
            )
            sc.fill = fill(C_GRIS_DARK)
            sc.font = Font(bold=True, size=9, color="F8C471", name="Calibri")
            sc.alignment = Alignment(horizontal="left", vertical="center")
            sc.border = border()
            ws.row_dimensions[current_row].height = 18
            current_row += 1

            # Detectar si es carrusel o post normal
            tipo_post = "Post Estándar"
            post_fb = art.get("Post_FB", "")
            post_ig = art.get("Post_IG", "")
            if "[CARRUSEL]" in post_fb or "[CARRUSEL]" in post_ig:
                tipo_post = "🎠 Carrusel"

            # Color de fila según silo
            silo = art.get("Silo", "")
            if "activaqr" in silo or "menu" in silo.lower():
                row_bg = "FEF9E7"  # Amarillo suave → ActivaQR
            elif "posicionamiento" in silo:
                row_bg = "EBF5FB"  # Azul suave → SEO Local
            else:
                row_bg = "F9FAFB"  # Blanco roto → general

            data = [
                f"S{sem_num}",
                art.get("ID", ""),
                art.get("Título", ""),
                silo,
                tipo_post,
                post_fb,
                post_ig,
                art.get("Post_LI", ""),
                art.get("Post_X", ""),
                art.get("Thumbnail", ""),
                art.get("URL_Final", ""),
            ]

            for col_idx, value in enumerate(data, start=1):
                c = ws.cell(row=current_row, column=col_idx, value=value)
                c.fill = fill(row_bg)
                c.border = border()

                # Alineación y fuente según columna
                if col_idx <= 5:
                    c.font = Font(size=9, bold=(col_idx == 1), name="Calibri")
                    c.alignment = center() if col_idx in (1, 2, 5) else left()
                else:
                    # Columnas de copy: fondo de color de red suave
                    red_bgs = {
                        6: "EBF5FB",  # FB suave
                        7: "F3E5F5",  # IG suave
                        8: "E8F4F8",  # LI suave
                        9: "F5F5F5",  # X suave
                    }
                    if col_idx in red_bgs:
                        c.fill = fill(red_bgs[col_idx])
                    c.font = Font(size=8.5, name="Calibri")
                    c.alignment = left()

            ws.row_dimensions[current_row].height = 90
            current_row += 1

    # Congelar encabezados
    ws.freeze_panes = "A3"

# ─── HOJA ÍNDICE ───────────────────────────────────────────────────────────
idx = wb.create_sheet(title="ÍNDICE", index=0)
idx.sheet_view.showGridLines = False

for col, w in [(1,25),(2,55),(3,20),(4,18)]:
    idx.column_dimensions[get_column_letter(col)].width = w

idx.merge_cells("A1:D1")
t = idx.cell(row=1, column=1, value="📋 CALENDARIO ESTRATÉGICO DE PUBLICACIONES RRSS 2026")
t.fill = fill(C_ROJO); t.font = Font(bold=True, size=13, color=C_BLANCO, name="Calibri")
t.alignment = center(); idx.row_dimensions[1].height = 26

idx.merge_cells("A2:D2")
s = idx.cell(row=2, column=1, value="César Reyes Jaramillo  ·  24 Artículos  ·  4 Redes Sociales  ·  Mes 1 · 2026")
s.fill = fill(C_GRIS_DARK); s.font = Font(size=10, color="F8C471", italic=True, name="Calibri")
s.alignment = center(); idx.row_dimensions[2].height = 20

for col_i, h in enumerate(["Día", "Contenido Programado", "# de Artículos", "Estado"], start=1):
    hdr = idx.cell(row=3, column=col_i, value=h)
    hdr.fill = fill(C_GRIS_DARK); hdr.font = Font(bold=True, size=9, color=C_BLANCO, name="Calibri")
    hdr.alignment = center(); hdr.border = border()
idx.row_dimensions[3].height = 22

dia_info = [
    ("LUNES",     "Contenido variado — Descubrimiento / Frustración / Comercial"),
    ("MARTES",    "Contenido variado — Educación / Solución / Transaccional"),
    ("MIÉRCOLES", "Análisis y Carruseles — Profundidad técnica y pilar SEO"),
    ("JUEVES",    "Micro-tips y Confianza — Formatos cortos, alto valor"),
    ("VIERNES",   "Conversión directa — CTA fuerte, caso real, urgencia"),
    ("SÁBADO",    "Carruseles y Contenido Visual — ActivaQR focus"),
    ("DOMINGO",   "🌿 Descanso — Sin publicaciones programadas"),
]

colors_idx = ["1A5276","1E8449","7B6000","7D3C98","C0392B","D68910","717D7E"]
for i, (dia, desc) in enumerate(dia_info, start=4):
    n_arts = len(dia_map.get(dia, []))
    idx.cell(row=i, column=1, value=dia).fill = fill(colors_idx[i-4])
    idx.cell(row=i, column=1).font = Font(bold=True, size=10, color=C_BLANCO, name="Calibri")
    idx.cell(row=i, column=1).alignment = center(); idx.cell(row=i, column=1).border = border()

    idx.cell(row=i, column=2, value=desc).font = Font(size=9, name="Calibri")
    idx.cell(row=i, column=2).alignment = left(); idx.cell(row=i, column=2).border = border()

    idx.cell(row=i, column=3, value=f"{n_arts} artículo(s)").font = Font(size=9, bold=True, name="Calibri")
    idx.cell(row=i, column=3).alignment = center(); idx.cell(row=i, column=3).border = border()

    estado = "⏳ Planificado" if n_arts > 0 else "🌿 Libre"
    idx.cell(row=i, column=4, value=estado).font = Font(size=9, bold=True,
        color="E74C3C" if n_arts > 0 else "1E8449", name="Calibri")
    idx.cell(row=i, column=4).alignment = center(); idx.cell(row=i, column=4).border = border()
    idx.row_dimensions[i].height = 20

# Leyenda de colores
idx.merge_cells("A12:D12")
ley = idx.cell(row=12, column=1, value="🎨 LEYENDA DE COLORES")
ley.fill = fill("ECF0F1"); ley.font = Font(bold=True, size=10, name="Calibri")
ley.alignment = center(); idx.row_dimensions[12].height = 20

leyenda = [
    (C_FB,    "Azul       → Copy de Facebook"),
    (C_IG,    "Púrpura    → Copy de Instagram"),
    (C_LI,    "Azul LI    → Copy de LinkedIn"),
    (C_X,     "Negro/Gris → Copy de X (Twitter)"),
    ("FEF9E7", "Amarillo   → Artículos de ActivaQR (menús, tarjetas)"),
    ("EBF5FB", "Azul suave → Artículos de SEO Local / Posicionamiento"),
    ("F9FAFB", "Blanco     → Artículos generales de Marketing"),
]
for i, (col_hex, desc) in enumerate(leyenda, start=13):
    idx.cell(row=i, column=1).fill = fill(col_hex)
    idx.cell(row=i, column=1).border = border()
    idx.cell(row=i, column=2, value=desc).font = Font(size=9, name="Calibri")
    idx.merge_cells(start_row=i, start_column=2, end_row=i, end_column=4)
    idx.cell(row=i, column=2).alignment = left(); idx.cell(row=i, column=2).border = border()
    idx.row_dimensions[i].height = 16

# ─── GUARDAR ───────────────────────────────────────────────────────────────
output = r"C:\Users\Cesar\Documents\GRUPO EMPRESARIAL REYES\PROYECTOS\RRSS_objetivo\CALENDARIO_POSTS_2026.xlsx"
wb.save(output)
print(f"✅ Excel generado:\n{output}")
print(f"📊 Total artículos procesados: {len(articles)}")
print(f"📅 Hojas creadas: {[ws.title for ws in wb.worksheets]}")
